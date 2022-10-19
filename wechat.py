import openpyxl as pxl
import pandas as pd
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
font = Font(name='微软雅黑', size=50, bold=False, italic=False, vertAlign=None,
                                        underline='none', strike=False, color='FFFF0000')



wb=pxl.load_workbook('最终的统计后.xlsx')

# rows = [3 4 5 1 2]


# sheet=wb.worksheets[0]

# sheet['A1'].value='1'
# sheet.columns('english').font=font
# sheet.cell('A2:A10').font=font

# column = sheet.column_dimensions['A']
# column.font = font

# for row in sheet:
#     for cell in row:
#         if cell.coordinate[0] == 'A':
#             print(sheet[cell.coordinate].value)

# italic24Font = Font( size = 24, italic = True, color='00FF0000' )
# column = sheet.column_dimensions['B']
# column.font = italic24Font

# sheet['A1'] = 'Hello world'
# sheet['A1'].font = italic24Font

# for row in sheet.iter_rows():
# 	for cell in row:
# 		print(cell.coordinate, cell.value)
# sheet['商品名称'].value='1111111'
# sheet.font=font
# # sheet.cell(1,1).font=font

# for row in sheet.iter_rows('1'):
# 	print(cell.coordinate, cell.value)

ws1 = wb.create_sheet("Sheet1")
ws1.title = "Title1"
ws2 = wb.create_sheet("Sheet2")
ws2.title = "Title2"

wb.save(filename="最终的统计后.xlsx")
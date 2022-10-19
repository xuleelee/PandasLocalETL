import openpyxl as pxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting import Rule
from openpyxl.formatting.rule import DataBarRule
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule

def my_yanse():
    wb=pxl.load_workbook('最终的统计后.xlsx')
    if '面包带单号' in wb.sheetnames:
        ws = wb['面包带单号']
        red_text = Font(color="9C0006")
        red_fill = PatternFill(bgColor="FFC7CE")
        dxf = DifferentialStyle(font=red_text, fill=red_fill)
        # rule = Rule(type="containsText", operator="containsText", text="highlight", dxf=dxf)
        # rule.formula = ['NOT(ISERROR(SEARCH("highlight",A1)))']
        rule = Rule(type="duplicateValues", dxf=dxf, stopIfTrue=None)
        ws.conditional_formatting.add('B1:B200', rule)
        wb.save("最终的统计后.xlsx")
    else:
        print('今天没有面包')

my_yanse()




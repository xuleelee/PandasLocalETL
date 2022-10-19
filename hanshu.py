import pandas as pd
import numpy as np
import openpyxl as pxl

from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting import Rule
from openpyxl.formatting.rule import DataBarRule
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
list = []
list.append('已订购:')
list_not_order = []
list_not_order.append('无订购:')



def my_chuju():
    milk = pd.read_csv('D:/TestForPandas/1.csv')



    data = milk.loc[milk['商品名称'].str.contains('✦')]
    if not data.empty:
        df_p = data.pivot_table(index='商品名称',  # 透视的行，分组依据
                                values='商品数量',  # 值
                                aggfunc='sum',  # 聚合函数
                                margins=True,
margins_name='合计'
                                )
        # print(data)
        df_p['商品数量'] = '共计' + df_p['商品数量'].astype('str') + '份'


        filename = '最终的统计后.xlsx'
        excel_book = pxl.load_workbook(filename)

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Your loaded workbook is set as the "base of work"
            writer.book = excel_book

            # Loop through the existing worksheets in the workbook and map each title to\
            # the corresponding worksheet (that is, a dictionary where the keys are the\
            # existing worksheets' names and the values are the actual worksheets)
            writer.sheets = {worksheet.title: worksheet for worksheet in excel_book.worksheets}

            # Write the new data to the file without overwriting what already exists
            df_p.to_excel(writer, '品牌厨具', index=True)

            # Save the file
            writer.save()
            list.append('品牌厨具')
    else:
        list_not_order.append('品牌厨具')
def my_dalei():
    milk = pd.read_csv('D:/TestForPandas/1.csv')

    data = milk.loc[milk['商品名称'].str.contains('※')]
    if not data.empty:
        df_p = data.pivot_table(index='商品名称',  # 透视的行，分组依据
                                values='商品数量',  # 值
                                aggfunc='sum',  # 聚合函数
                                margins=True,
margins_name='合计'
                                )
        # print(data)
        df_p['商品数量'] = '共计' + df_p['商品数量'].astype('str') + '份'


        filename = '最终的统计后.xlsx'
        excel_book = pxl.load_workbook(filename)

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Your loaded workbook is set as the "base of work"
            writer.book = excel_book

            # Loop through the existing worksheets in the workbook and map each title to\
            # the corresponding worksheet (that is, a dictionary where the keys are the\
            # existing worksheets' names and the values are the actual worksheets)
            writer.sheets = {worksheet.title: worksheet for worksheet in excel_book.worksheets}

            # Write the new data to the file without overwriting what already exists
            df_p.to_excel(writer, '大磊', index=True)

            # Save the file
            writer.save()
            list.append('大磊')
    else:
        list_not_order.append('大磊')
def my_meidi():
    milk = pd.read_csv('D:/TestForPandas/1.csv')

    data = milk.loc[milk['商品名称'].str.contains('■')]
    if not data.empty:
        df_p = data.pivot_table(index='商品名称',  # 透视的行，分组依据
                                values='商品数量',  # 值
                                aggfunc='sum',  # 聚合函数
                                margins=True,
margins_name='合计'
                                )
        # print(data)
        df_p['商品数量'] = '共计' + df_p['商品数量'].astype('str') + '份'


        filename = '最终的统计后.xlsx'
        excel_book = pxl.load_workbook(filename)

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Your loaded workbook is set as the "base of work"
            writer.book = excel_book

            # Loop through the existing worksheets in the workbook and map each title to\
            # the corresponding worksheet (that is, a dictionary where the keys are the\
            # existing worksheets' names and the values are the actual worksheets)
            writer.sheets = {worksheet.title: worksheet for worksheet in excel_book.worksheets}

            # Write the new data to the file without overwriting what already exists
            df_p.to_excel(writer, '美的', index=True)

            # Save the file
            writer.save()
            list.append('美的')
    else:
        list_not_order.append('美的')

def my_yan():
    milk = pd.read_csv('D:/TestForPandas/1.csv')

    data = milk.loc[milk['商品名称'].str.contains('☀')]
    if not data.empty:
        df_p = data.pivot_table(index='商品名称',  # 透视的行，分组依据
                                values='商品数量',  # 值
                                aggfunc='sum',  # 聚合函数
                                margins=True,
margins_name='合计'
                                )
        # print(data)
        df_p['商品数量'] = '共计' + df_p['商品数量'].astype('str') + '份'


        filename = '最终的统计后.xlsx'
        excel_book = pxl.load_workbook(filename)

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Your loaded workbook is set as the "base of work"
            writer.book = excel_book

            # Loop through the existing worksheets in the workbook and map each title to\
            # the corresponding worksheet (that is, a dictionary where the keys are the\
            # existing worksheets' names and the values are the actual worksheets)
            writer.sheets = {worksheet.title: worksheet for worksheet in excel_book.worksheets}

            # Write the new data to the file without overwriting what already exists
            df_p.to_excel(writer, '烟', index=True)

            # Save the file
            writer.save()
            list.append('烟')
    else:
        list_not_order.append('烟')

def my_zahuo():
    milk = pd.read_csv('D:/TestForPandas/1.csv')

    data = milk.loc[milk['商品名称'].str.contains('✡')]
    if not data.empty:
        df_p = data.pivot_table(index='商品名称',  # 透视的行，分组依据
                                values='商品数量',  # 值
                                aggfunc='sum',  # 聚合函数
                                margins=True,
margins_name='合计'
                                )
        # print(data)
        df_p['商品数量'] = '共计' + df_p['商品数量'].astype('str') + '份'


        filename = '最终的统计后.xlsx'
        excel_book = pxl.load_workbook(filename)

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Your loaded workbook is set as the "base of work"
            writer.book = excel_book

            # Loop through the existing worksheets in the workbook and map each title to\
            # the corresponding worksheet (that is, a dictionary where the keys are the\
            # existing worksheets' names and the values are the actual worksheets)
            writer.sheets = {worksheet.title: worksheet for worksheet in excel_book.worksheets}

            # Write the new data to the file without overwriting what already exists
            df_p.to_excel(writer, '刘杂货', index=True)

            # Save the file
            writer.save()
            list.append('杂货')
    else:
        list_not_order.append('杂货')

def highlight_all():

    return ['background: yellow' if x.name in rows else '' for i in x]

def my_mianbao():




    original_table = pd.read_csv('D:/TestForPandas/1.csv')
    information_table = pd.read_excel('D:/TestForPandas/mianbao.xlsx', sheet_name='Sheet1')



    result = original_table.merge(information_table, on="商品名称")
    mianbao = result[['订单号','商品名称','商品数量']].copy()

    mianbao.sort_values("订单号", ascending=[False],inplace=True)
    mianbao.sort_values("订单号",  inplace=True)
    mianbao['订单号']= mianbao['订单号'].astype(str)
    mianbao['订单号'] = mianbao['订单号'].str[13:18]
    mianbao[['商品名称', '商品数量']] = mianbao[['商品数量','商品名称']]
    # print(mianbao['订单号'])
    # chongfu = mianbao['订单号']
    # print(chongfu)

    # rows_series = mianbao[['订单号']].duplicated(keep=False)
    # print(rows_series)
    # rows = rows_series[rows_series].index.values
    # print(rows)
    # yanse = mianbao.style.apply(lambda x: ['background: yellow' if x.name in rows else '' for i in x], axis=1)
    # yanse = mianbao.style.apply(lambda x: ['background: yellow' if x.name in rows else '' for i in x])
    # mianbao.style.format({'订单号':'{:.2%}'})
    # # print(mianbao['Duplicate'])
    # print(mianbao)


    if not result.empty:
        vege_pivot = result.pivot_table(index='商品名称',  # 透视的行，分组依据
                                        values='商品数量',  # 值
                                        aggfunc='sum',  # 聚合函数
                                        margins=True,
                                        margins_name='合计'
                                        )


        vege_pivot['商品数量'] = '共计' + vege_pivot['商品数量'].astype('str') + '份'

        filename = '最终的统计后.xlsx'
        excel_book = pxl.load_workbook(filename)

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Your loaded workbook is set as the "base of work"
            writer.book = excel_book

            # Loop through the existing worksheets in the workbook and map each title to\
            # the corresponding worksheet (that is, a dictionary where the keys are the\
            # existing worksheets' names and the values are the actual worksheets)
            writer.sheets = {worksheet.title: worksheet for worksheet in excel_book.worksheets}

            # Write the new data to the file without overwriting what already exists
            vege_pivot.to_excel(writer, '面包总数', index=True)
            mianbao.to_excel(writer, '面包带单号', index=True)
            # Save the file
            writer.save()
            list.append('面包')
    else:
        list_not_order.append('面包')


def my_ji():




    original_table = pd.read_csv('D:/TestForPandas/1.csv')
    information_table = pd.read_excel('D:/TestForPandas/ji.xlsx', sheet_name='Sheet1')



    result = original_table.merge(information_table, on="商品名称")
    if not result.empty:
        result['商品名称'] = result['商品名称'].str.replace('(', ' ')
        result['商品名称'] = result['商品名称'].str.replace('注: 以实际称重结算', ' ')
        result['商品名称'] = result['商品名称'].str.replace(')', ' ')
        result['商品名称'] = result['商品名称'].str.replace('）', ' ')

        vege_pivot = result.pivot_table(index='商品名称',  # 透视的行，分组依据
                                        values='商品数量',  # 值
                                        aggfunc='sum',  # 聚合函数
                                        margins=True,
margins_name='合计'
                                        )

        # print(vege_pivot)
        vege_pivot['商品数量'] = '共计' + vege_pivot['商品数量'].astype('str') + '份'

        filename = '最终的统计后.xlsx'
        excel_book = pxl.load_workbook(filename)

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Your loaded workbook is set as the "base of work"
            writer.book = excel_book

            # Loop through the existing worksheets in the workbook and map each title to\
            # the corresponding worksheet (that is, a dictionary where the keys are the\
            # existing worksheets' names and the values are the actual worksheets)
            writer.sheets = {worksheet.title: worksheet for worksheet in excel_book.worksheets}

            # Write the new data to the file without overwriting what already exists
            vege_pivot.to_excel(writer, '鸡', index=True)

            # Save the file
            writer.save()
            list.append('鸡')

    else:
        list_not_order.append('鸡')
def my_rou():




    original_table = pd.read_csv('D:/TestForPandas/1.csv')
    information_table = pd.read_excel('D:/TestForPandas/rou.xlsx', sheet_name='Sheet1')



    result = original_table.merge(information_table, on="商品名称")
    if not result.empty:
        result['商品名称'] = result['商品名称'].str.replace('(', ' ')
        result['商品名称'] = result['商品名称'].str.replace('注: 以实际称重结算', ' ')
        result['商品名称'] = result['商品名称'].str.replace(')', ' ')
        result['商品名称'] = result['商品名称'].str.replace('）', ' ')

        vege_pivot = result.pivot_table(index='商品名称',  # 透视的行，分组依据
                                        values='商品数量',  # 值
                                        aggfunc='sum',  # 聚合函数
                                        margins=True,
                                        margins_name='合计'
                                        )
        # print(vege_pivot)
        # vege_pivot['商品名称'] = vege_pivot['商品名称'].str.replace('All', '合计')
        # print(vege_pivot)
        vege_pivot['商品数量'] = '共计' + vege_pivot['商品数量'].astype('str') + '份'

        filename = '最终的统计后.xlsx'
        excel_book = pxl.load_workbook(filename)

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Your loaded workbook is set as the "base of work"
            writer.book = excel_book

            # Loop through the existing worksheets in the workbook and map each title to\
            # the corresponding worksheet (that is, a dictionary where the keys are the\
            # existing worksheets' names and the values are the actual worksheets)
            writer.sheets = {worksheet.title: worksheet for worksheet in excel_book.worksheets}

            # Write the new data to the file without overwriting what already exists
            vege_pivot.to_excel(writer, '肉', index=True)

            # Save the file
            writer.save()
            list.append('肉')
    else:
        list_not_order.append('肉')




def my_laoshanghai():
    milk = pd.read_csv('D:/TestForPandas/1.csv')

    data = milk.loc[milk['商品名称'].str.contains('❥')]
    if not data.empty:
        df_p = data.pivot_table(index='商品名称',  # 透视的行，分组依据
                                values='商品数量',  # 值
                                aggfunc='sum',  # 聚合函数
                                margins=True,
margins_name='合计'
                                )
        # print(data)
        df_p['商品数量'] = '共计' + df_p['商品数量'].astype('str') + '份'


        filename = '最终的统计后.xlsx'
        excel_book = pxl.load_workbook(filename)

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Your loaded workbook is set as the "base of work"
            writer.book = excel_book

            # Loop through the existing worksheets in the workbook and map each title to\
            # the corresponding worksheet (that is, a dictionary where the keys are the\
            # existing worksheets' names and the values are the actual worksheets)
            writer.sheets = {worksheet.title: worksheet for worksheet in excel_book.worksheets}

            # Write the new data to the file without overwriting what already exists
            df_p.to_excel(writer, '老上海', index=True)

            # Save the file
            writer.save()
            list.append('老上海')
    else:
        list_not_order.append('老上海')

def my_milktea():
    milk = pd.read_csv('D:/TestForPandas/1.csv')
    information_table = pd.read_csv('D:/TestForPandas/2.csv')
    # print(df.head(5))
    data = milk.loc[milk['商品名称'].str.contains('☺')]
    if not data.empty:
        df_p = data.pivot_table(index='商品名称',  # 透视的行，分组依据
                                values='商品数量',  # 值
                                aggfunc='sum',  # 聚合函数
                                margins=True,
margins_name='合计'
                                )
        # print(data)
        df_p['商品数量'] = '共计' + df_p['商品数量'].astype('str') + '份'


        filename = '最终的统计后.xlsx'
        excel_book = pxl.load_workbook(filename)

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Your loaded workbook is set as the "base of work"
            writer.book = excel_book

            # Loop through the existing worksheets in the workbook and map each title to\
            # the corresponding worksheet (that is, a dictionary where the keys are the\
            # existing worksheets' names and the values are the actual worksheets)
            writer.sheets = {worksheet.title: worksheet for worksheet in excel_book.worksheets}

            # Write the new data to the file without overwriting what already exists
            df_p.to_excel(writer, '奶茶', index=True)

            # Save the file
            writer.save()
            list.append('奶茶')
    else:
        list_not_order.append('奶茶')

def my_haixian():
    milk = pd.read_csv('D:/TestForPandas/1.csv')
    information_table = pd.read_csv('D:/TestForPandas/2.csv')
    # print(df.head(5))
    data = milk.loc[milk['商品名称'].str.contains('ゆ')]
    if not data.empty:
        df_p = data.pivot_table(index='商品名称',  # 透视的行，分组依据
                                values='商品数量',  # 值
                                aggfunc='sum',  # 聚合函数
                                margins=True,
margins_name='合计'
                                )
        # print(data)
        df_p['商品数量'] = '共计' + df_p['商品数量'].astype('str') + '份'


        filename = '最终的统计后.xlsx'
        excel_book = pxl.load_workbook(filename)

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Your loaded workbook is set as the "base of work"
            writer.book = excel_book

            # Loop through the existing worksheets in the workbook and map each title to\
            # the corresponding worksheet (that is, a dictionary where the keys are the\
            # existing worksheets' names and the values are the actual worksheets)
            writer.sheets = {worksheet.title: worksheet for worksheet in excel_book.worksheets}

            # Write the new data to the file without overwriting what already exists
            df_p.to_excel(writer, '海鲜', index=True)

            # Save the file
            writer.save()
            list.append('海鲜')
    else:
        list_not_order.append('海鲜')
def my_baiwei():
    milk = pd.read_csv('D:/TestForPandas/1.csv')
    information_table = pd.read_csv('D:/TestForPandas/2.csv')
    # print(df.head(5))
    data = milk.loc[milk['商品名称'].str.contains('↑')]
    if not data.empty:
        df_p = data.pivot_table(index='商品名称',  # 透视的行，分组依据
                                values='商品数量',  # 值
                                aggfunc='sum',  # 聚合函数
                                margins=True,
margins_name='合计'
                                )
        # print(data)
        df_p['商品数量'] = '共计' + df_p['商品数量'].astype('str') + '份'


        filename = '最终的统计后.xlsx'
        excel_book = pxl.load_workbook(filename)

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Your loaded workbook is set as the "base of work"
            writer.book = excel_book

            # Loop through the existing worksheets in the workbook and map each title to\
            # the corresponding worksheet (that is, a dictionary where the keys are the\
            # existing worksheets' names and the values are the actual worksheets)
            writer.sheets = {worksheet.title: worksheet for worksheet in excel_book.worksheets}

            # Write the new data to the file without overwriting what already exists
            df_p.to_excel(writer, '百味捞', index=True)

            # Save the file
            writer.save()
            list.append('百味捞')
    else:
        list_not_order.append('百味捞')

def my_gebi():
    milk = pd.read_csv('D:/TestForPandas/1.csv')
    information_table = pd.read_csv('D:/TestForPandas/2.csv')
    # print(df.head(5))
    data = milk.loc[milk['商品名称'].str.contains('↣')]
    if not data.empty:
        df_p = data.pivot_table(index='商品名称',  # 透视的行，分组依据
                                values='商品数量',  # 值
                                aggfunc='sum',  # 聚合函数
                                margins=True,
margins_name='合计'
                                )
        # print(data)
        df_p['商品数量'] = '共计' + df_p['商品数量'].astype('str') + '份'


        filename = '最终的统计后.xlsx'
        excel_book = pxl.load_workbook(filename)

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Your loaded workbook is set as the "base of work"
            writer.book = excel_book

            # Loop through the existing worksheets in the workbook and map each title to\
            # the corresponding worksheet (that is, a dictionary where the keys are the\
            # existing worksheets' names and the values are the actual worksheets)
            writer.sheets = {worksheet.title: worksheet for worksheet in excel_book.worksheets}

            # Write the new data to the file without overwriting what already exists
            df_p.to_excel(writer, '隔壁海鲜', index=True)

            # Save the file
            writer.save()
            list.append('隔壁海鲜')
    else:
        list_not_order.append('隔壁海鲜')


def my_heniu():
    milk = pd.read_csv('D:/TestForPandas/1.csv')
    information_table = pd.read_csv('D:/TestForPandas/2.csv')
    # print(df.head(5))
    data = milk.loc[milk['商品名称'].str.contains('♥')]
    if not data.empty:
        df_p = data.pivot_table(index='商品名称',  # 透视的行，分组依据
                                values='商品数量',  # 值
                                aggfunc='sum',  # 聚合函数
                                margins=True,
margins_name='合计'
                                )
        # print(data)
        df_p['商品数量'] = '共计' + df_p['商品数量'].astype('str') + '份'


        filename = '最终的统计后.xlsx'
        excel_book = pxl.load_workbook(filename)

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Your loaded workbook is set as the "base of work"
            writer.book = excel_book

            # Loop through the existing worksheets in the workbook and map each title to\
            # the corresponding worksheet (that is, a dictionary where the keys are the\
            # existing worksheets' names and the values are the actual worksheets)
            writer.sheets = {worksheet.title: worksheet for worksheet in excel_book.worksheets}

            # Write the new data to the file without overwriting what already exists
            df_p.to_excel(writer, '和牛', index=True)

            # Save the file
            writer.save()
            list.append('和牛')
    else:
        list_not_order.append('和牛')

def my_aojia():
    milk = pd.read_csv('D:/TestForPandas/1.csv')
    information_table = pd.read_csv('D:/TestForPandas/2.csv')
    # print(df.head(5))
    data = milk.loc[milk['商品名称'].str.contains('❆')]
    if not data.empty:
        df_p = data.pivot_table(index='商品名称',  # 透视的行，分组依据
                                values='商品数量',  # 值
                                aggfunc='sum',  # 聚合函数
                                margins=True,
margins_name='合计'
                                )

        df_p['商品数量'] = '共计' + df_p['商品数量'].astype('str') + '份'


        filename = '最终的统计后.xlsx'
        excel_book = pxl.load_workbook(filename)

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Your loaded workbook is set as the "base of work"
            writer.book = excel_book

            # Loop through the existing worksheets in the workbook and map each title to\
            # the corresponding worksheet (that is, a dictionary where the keys are the\
            # existing worksheets' names and the values are the actual worksheets)
            writer.sheets = {worksheet.title: worksheet for worksheet in excel_book.worksheets}

            # Write the new data to the file without overwriting what already exists
            df_p.to_excel(writer, '澳佳', index=True)

            # Save the file
            writer.save()
            list.append('澳佳')
    else:
        list_not_order.append('澳佳')

def my_shidai():
    milk = pd.read_csv('D:/TestForPandas/1.csv')
    information_table = pd.read_csv('D:/TestForPandas/2.csv')
    # print(df.head(5))
    data = milk.loc[milk['商品名称'].str.contains('◆')]
    if not data.empty:
        df_p = data.pivot_table(index='商品名称',  # 透视的行，分组依据
                                values='商品数量',  # 值
                                aggfunc='sum',  # 聚合函数
                                margins=True,
margins_name='合计'
                                )
        # print(data)
        df_p['商品数量'] = '共计' + df_p['商品数量'].astype('str') + '份'

        filename = '最终的统计后.xlsx'
        excel_book = pxl.load_workbook(filename)

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Your loaded workbook is set as the "base of work"
            writer.book = excel_book

            # Loop through the existing worksheets in the workbook and map each title to\
            # the corresponding worksheet (that is, a dictionary where the keys are the\
            # existing worksheets' names and the values are the actual worksheets)
            writer.sheets = {worksheet.title: worksheet for worksheet in excel_book.worksheets}

            # Write the new data to the file without overwriting what already exists
            df_p.to_excel(writer, '时代', index=True)

            # Save the file
            writer.save()
            list.append('时代')
    else:
        list_not_order.append('时代')

def my_guanya():

    milk = pd.read_csv('D:/TestForPandas/1.csv')
    information_table = pd.read_csv('D:/TestForPandas/2.csv')
    # print(df.head(5))
    data = milk.loc[milk['商品名称'].str.contains('♣')]
    if not data.empty:
        df_p = data.pivot_table(index='商品名称',  # 透视的行，分组依据
                                values='商品数量',  # 值
                                aggfunc='sum',  # 聚合函数
                                margins=True,
margins_name='合计'
                                )
        # print(data)
        df_p['商品数量'] = '共计' + df_p['商品数量'].astype('str') + '份'

        filename = '最终的统计后.xlsx'
        excel_book = pxl.load_workbook(filename)

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Your loaded workbook is set as the "base of work"
            writer.book = excel_book

            # Loop through the existing worksheets in the workbook and map each title to\
            # the corresponding worksheet (that is, a dictionary where the keys are the\
            # existing worksheets' names and the values are the actual worksheets)
            writer.sheets = {worksheet.title: worksheet for worksheet in excel_book.worksheets}

            # Write the new data to the file without overwriting what already exists
            df_p.to_excel(writer, '冠亚', index=True)

            # Save the file
            writer.save()
            list.append('冠亚')
    else:
        list_not_order.append('冠亚')
def my_xinchangfa():
    milk = pd.read_csv('D:/TestForPandas/1.csv')
    information_table = pd.read_csv('D:/TestForPandas/2.csv')
    # print(df.head(5))
    data = milk.loc[milk['商品名称'].str.contains('✿')]
    if not data.empty:
        df_p = data.pivot_table(index='商品名称',  # 透视的行，分组依据
                                values='商品数量',  # 值
                                aggfunc='sum',  # 聚合函数
                                margins=True,
margins_name='合计'
                                )
        # print(data)
        df_p['商品数量'] = '共计' + df_p['商品数量'].astype('str') + '份'

        filename = '最终的统计后.xlsx'
        excel_book = pxl.load_workbook(filename)

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Your loaded workbook is set as the "base of work"
            writer.book = excel_book

            # Loop through the existing worksheets in the workbook and map each title to\
            # the corresponding worksheet (that is, a dictionary where the keys are the\
            # existing worksheets' names and the values are the actual worksheets)
            writer.sheets = {worksheet.title: worksheet for worksheet in excel_book.worksheets}

            # Write the new data to the file without overwriting what already exists
            df_p.to_excel(writer, '新长发', index=True)

            # Save the file
            writer.save()
            list.append('新长发')
    else:
        list_not_order.append('新长发')

def my_niunai():
    milk = pd.read_csv('D:/TestForPandas/1.csv')
    information_table = pd.read_csv('D:/TestForPandas/2.csv')
    # print(df.head(5))
    data = milk.loc[milk['商品名称'].str.contains('❄')]
    if not data.empty:
        df_p = data.pivot_table(index='商品名称',  # 透视的行，分组依据
                                values='商品数量',  # 值
                                aggfunc='sum',  # 聚合函数
                                margins=True,
margins_name='合计'
                                )
        # print(data)
        df_p['商品数量'] = '共计' + df_p['商品数量'].astype('str') + '份'

        filename = '最终的统计后.xlsx'
        excel_book = pxl.load_workbook(filename)

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Your loaded workbook is set as the "base of work"
            writer.book = excel_book

            # Loop through the existing worksheets in the workbook and map each title to\
            # the corresponding worksheet (that is, a dictionary where the keys are the\
            # existing worksheets' names and the values are the actual worksheets)
            writer.sheets = {worksheet.title: worksheet for worksheet in excel_book.worksheets}

            # Write the new data to the file without overwriting what already exists
            df_p.to_excel(writer, '牛奶', index=True)

            # Save the file
            writer.save()
            list.append('牛奶')
    else:
        list_not_order.append('牛奶')

def my_abs():




    original_table = pd.read_csv('D:/TestForPandas/1.csv')
    information_table = pd.read_excel('D:/TestForPandas/information.xlsx', sheet_name='vege')

    original_table.loc[(original_table['商品名称'].str.contains('上海青')), ['商品数量']] = original_table.loc[(original_table[
                                                                                                         '商品名称'].str.contains(
        '上海青')), ['商品数量']] * 3
    original_table.loc[(original_table['商品名称'].str.contains('青菜心')), ['商品数量']] = original_table.loc[(original_table[
                                                                                                         '商品名称'].str.contains(
        '青菜心')), ['商品数量']] * 2
    original_table.loc[(original_table['商品名称'].str.contains('芥蓝')), ['商品数量']] = original_table.loc[(original_table[
                                                                                                        '商品名称'].str.contains(
        '芥蓝')), ['商品数量']] * 2
    original_table.loc[(original_table['商品名称'].str.contains('樱桃萝卜')), ['商品数量']] = original_table.loc[(original_table[
                                                                                                          '商品名称'].str.contains(
        '樱桃萝卜')), ['商品数量']] * 3


    result = original_table.merge(information_table, on="商品名称")
    # result = pd.merge(original_table,information_table.loc[:,['商品名称','英语名字']],how='right',on = '商品名称')
    # print(result)

    vege_pivot = result.pivot_table(index='english',  # 透视的行，分组依据
                                    values='商品数量',  # 值
                                    aggfunc='sum',  # 聚合函数
                                    margins=True,
margins_name='合计'
                                    )


    final_df = result = vege_pivot.merge(information_table, on="english")



    # without_duplication = final_df.drop_duplicates(['english'],keep='last',inplace=True)
    final_test = final_df.duplicated(subset=['english'], keep='last')


    finally_test = final_df.drop_duplicates(subset=['english'], keep='last')



    writer = pd.ExcelWriter('最终的统计后.xlsx')
    finally_test.to_excel(writer, sheet_name='蔬菜')
    writer.save()
    list.append('蔬菜')

my_abs()
my_guanya()
my_shidai()
my_aojia()
my_niunai()
my_gebi()
my_heniu()
my_xinchangfa()
my_baiwei()
my_haixian()
my_milktea()
my_laoshanghai()
my_rou()
my_ji()
my_mianbao()
my_zahuo()
my_yan()
my_meidi()
my_dalei()
my_chuju()



# print(list)
StrA = " ".join(list)
print(StrA)
# print('无订购:')
# print(list_not_order)
StrB = " ".join(list_not_order)
print(StrB)

def my_yanse():
    wb=pxl.load_workbook('最终的统计后.xlsx')
    if '面包带单号' in wb.sheetnames:
        ws = wb['面包带单号']
        red_text = Font(color="9C0006")
        red_fill = PatternFill(bgColor="FFC7CE")
        dxf = DifferentialStyle(font=red_text, fill=red_fill)
        # rule = Rule(type="containsText", operator="containsText", text="highlight", dxf=dxf)
        # rule.formula = ['NOT(ISERROR(SEARCH("highlight",A1)))']
        rule = Rule(type="duplicateValues", dxf=dxf, stopIfTrue=None)
        ws.conditional_formatting.add('B1:B200', rule)
        wb.save("最终的统计后.xlsx")
    else:
        print('今天没有面包')

my_yanse()

# wb=pxl.load_workbook('最终的统计后.xlsx')
# sheet=wb['冠亚']
#
# sheet['A1'].value='1111111'
# # sheet.font=font
# # # sheet.cell(1,1).font=font

